#!/usr/bin/env python3
"""
Simple standalone API for VPS deployment
Minimal version with just essential functionality
"""

from fastapi import FastAPI, UploadFile, File, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any, AsyncGenerator
import httpx
import json
import os
import uuid
import aiofiles
from pathlib import Path
import logging
import base64
from PIL import Image
import PyPDF2
from docx import Document
import openpyxl
import asyncio
import time
from datetime import datetime
from contextlib import asynccontextmanager
import re

# Import monitoring, security, and logging
from monitoring import router as monitoring_router, monitoring_loop
from security import (
    SecurityManager, RateLimitMiddleware, SecurityHeadersMiddleware,
    get_optional_user, validate_api_key, secure_filename,
    log_security_event, monitor_security_events, security_manager
)
from logging_config import (
    logging_manager, error_tracker, performance_tracker,
    log_error, performance_monitor, error_handler, RequestLogger,
    get_logging_status, log_system_health
)
from performance_optimizer import performance_optimizer, LoadTester
from enhanced_model_router import enhanced_router
from token_manager import token_manager, ChunkingStrategy
from content_summarizer import ContentSummarizer

# Configure logging (enhanced logging is set up by logging_manager)
logger = logging.getLogger("autopicker.api")

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Application lifespan events"""
    # Startup
    logger.info("Starting production monitoring, security, and logging...")
    monitoring_task = asyncio.create_task(monitoring_loop())
    security_task = asyncio.create_task(monitor_security_events())
    health_logging_task = asyncio.create_task(log_system_health())
    
    yield
    
    # Shutdown
    logger.info("Shutting down monitoring, security, and logging...")
    monitoring_task.cancel()
    security_task.cancel()
    health_logging_task.cancel()
    try:
        await monitoring_task
        await security_task
        await health_logging_task
    except asyncio.CancelledError:
        pass

# Initialize FastAPI app
app = FastAPI(
    title="Multimodal LLM Platform API - Simple",
    description="Simplified API for VPS deployment",
    version="1.0.0",
    lifespan=lifespan
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Include monitoring router
app.include_router(monitoring_router)

# Add security and logging middleware
app.add_middleware(SecurityHeadersMiddleware)
app.add_middleware(RateLimitMiddleware)

# Initialize request logger
request_logger = RequestLogger()

# Configuration
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Initialize content summarizer
content_summarizer = ContentSummarizer(token_manager.token_counter)

# File processing functions
def process_pdf(file_path: Path) -> str:
    """Extract text from PDF file"""
    try:
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error processing PDF: {e}")
        return f"Error processing PDF: {str(e)}"

def process_docx(file_path: Path) -> str:
    """Extract text from DOCX file"""
    try:
        doc = Document(file_path)
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error processing DOCX: {e}")
        return f"Error processing DOCX: {str(e)}"

def process_excel(file_path: Path) -> str:
    """Extract text from Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_path)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"Error processing Excel: {e}")
        return f"Error processing Excel: {str(e)}"

def process_image(file_path: Path) -> Dict[str, Any]:
    """Process image file and return metadata"""
    try:
        with Image.open(file_path) as img:
            # Convert image to base64 for potential future use
            import io
            buffered = io.BytesIO()
            img.save(buffered, format=img.format or "PNG")
            img_base64 = base64.b64encode(buffered.getvalue()).decode()
            
            return {
                "type": "image",
                "size": img.size,
                "mode": img.mode,
                "format": img.format,
                "description": f"Image: {img.size[0]}x{img.size[1]} pixels, {img.mode} mode",
                "base64": img_base64[:100] + "..." if len(img_base64) > 100 else img_base64  # Truncated for preview
            }
    except Exception as e:
        logger.error(f"Error processing image: {e}")
        return {"type": "image", "error": str(e)}

def process_text(file_path: Path) -> str:
    """Process text file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()
    except Exception as e:
        logger.error(f"Error processing text file: {e}")
        return f"Error processing text file: {str(e)}"

async def process_audio(file_path: Path) -> Dict[str, Any]:
    """Process audio file and return transcription"""
    try:
        logger.info(f"Processing audio file: {file_path}")
        
        # Prepare file for Whisper API
        async with aiofiles.open(file_path, 'rb') as audio_file:
            files = {'audio': await audio_file.read()}
        
        # Send to Whisper service
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(
                "http://localhost:9002/asr",
                files={'audio': files['audio']},
                params={'output': 'json'}
            )
            
            if response.status_code == 200:
                result = response.json()
                transcription = result.get('text', '').strip()
                
                return {
                    "type": "audio",
                    "transcription": transcription,
                    "language": result.get('language', 'unknown'),
                    "duration": result.get('duration', 0),
                    "content": transcription  # For compatibility with text processing
                }
            else:
                logger.error(f"Whisper API error: {response.status_code} - {response.text}")
                return {
                    "type": "audio", 
                    "error": f"Transcription failed: {response.status_code}",
                    "fallback": f"Audio file ({file_path.suffix}) uploaded but transcription unavailable"
                }
                
    except Exception as e:
        logger.error(f"Error processing audio: {e}")
        return {
            "type": "audio", 
            "error": str(e),
            "fallback": f"Audio file ({file_path.suffix}) uploaded but processing failed"
        }

def get_file_content(file_path: Path) -> Dict[str, Any]:
    """Process any supported file type"""
    suffix = file_path.suffix.lower()
    
    if suffix == '.pdf':
        content = process_pdf(file_path)
        return {"type": "pdf", "content": content}
    elif suffix == '.docx':
        content = process_docx(file_path)
        return {"type": "docx", "content": content}
    elif suffix in ['.xlsx', '.xls']:
        content = process_excel(file_path)
        return {"type": "excel", "content": content}
    elif suffix in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
        content = process_image(file_path)
        return content
    elif suffix in ['.mp3', '.wav', '.m4a', '.ogg', '.flac']:
        # Audio files need async processing, return placeholder for now
        return {"type": "audio", "content": "Audio file uploaded, use transcription endpoint"}
    elif suffix in ['.txt', '.md', '.csv', '.json', '.py', '.js', '.html', '.css']:
        content = process_text(file_path)
        return {"type": "text", "content": content}
    else:
        return {"type": "unknown", "error": f"Unsupported file type: {suffix}"}

# Request/Response Models
class ChatMessage(BaseModel):
    role: str
    content: str

class ChatCompletionRequest(BaseModel):
    messages: List[ChatMessage]
    model: Optional[str] = "llama3.2-local"
    temperature: Optional[float] = 0.7
    max_tokens: Optional[int] = None
    stream: Optional[bool] = False

class FileUploadResponse(BaseModel):
    id: str
    filename: str
    original_filename: str
    size: int
    mime_type: str
    content_preview: Optional[str] = None
    file_type: Optional[str] = None

class MultimodalRequest(BaseModel):
    messages: List[ChatMessage]
    file_ids: Optional[List[str]] = []
    model: Optional[str] = "auto"  # Changed default to "auto" for smart routing
    temperature: Optional[float] = 0.7
    max_tokens: Optional[int] = None
    stream: Optional[bool] = False

# Note: Old ModelSelector class replaced with enhanced_model_router
# The enhanced router provides access to multiple models through OpenRouter and local Ollama

# Streaming helper functions
async def stream_ollama_response(payload: Dict[str, Any]) -> AsyncGenerator[str, None]:
    """Stream response from Ollama API"""
    try:
        # Add streaming to payload
        payload["stream"] = True
        
        async with httpx.AsyncClient(timeout=120.0) as client:
            async with client.stream("POST", "http://localhost:11434/api/chat", json=payload) as response:
                if response.status_code != 200:
                    logger.error(f"Ollama streaming error: {response.status_code}")
                    yield f"data: {json.dumps({'error': f'Ollama error: {response.status_code}'})}\n\n"
                    return
                
                async for chunk in response.aiter_text():
                    if chunk.strip():
                        try:
                            # Parse each JSON line from Ollama
                            for line in chunk.strip().split('\n'):
                                if line:
                                    ollama_response = json.loads(line)
                                    
                                    # Convert Ollama streaming format to OpenAI format
                                    if "message" in ollama_response and "content" in ollama_response["message"]:
                                        content = ollama_response["message"]["content"]
                                        
                                        openai_chunk = {
                                            "id": str(uuid.uuid4()),
                                            "object": "chat.completion.chunk",
                                            "model": payload.get("model", "llama3.2:1b"),
                                            "choices": [{
                                                "index": 0,
                                                "delta": {"content": content},
                                                "finish_reason": None
                                            }]
                                        }
                                        
                                        yield f"data: {json.dumps(openai_chunk)}\n\n"
                                    
                                    # Check if this is the final chunk
                                    if ollama_response.get("done", False):
                                        final_chunk = {
                                            "id": str(uuid.uuid4()),
                                            "object": "chat.completion.chunk",
                                            "model": payload.get("model", "llama3.2:1b"),
                                            "choices": [{
                                                "index": 0,
                                                "delta": {},
                                                "finish_reason": "stop"
                                            }]
                                        }
                                        yield f"data: {json.dumps(final_chunk)}\n\n"
                                        yield "data: [DONE]\n\n"
                                        return
                                        
                        except json.JSONDecodeError as e:
                            logger.error(f"JSON decode error in streaming: {e}")
                            continue
                            
    except Exception as e:
        logger.error(f"Streaming error: {e}")
        yield f"data: {json.dumps({'error': str(e)})}\n\n"

# Enhanced streaming function using the new router
async def stream_enhanced_response(model_id: str, messages: List[Dict], **kwargs) -> AsyncGenerator[str, None]:
    """Stream response using enhanced model router"""
    try:
        async for chunk in enhanced_router.make_api_call(model_id, messages, stream=True, **kwargs):
            yield f"data: {chunk}\n\n"
        
        yield "data: [DONE]\n\n"
        
    except Exception as e:
        logger.error(f"Enhanced router streaming error: {e}")
        yield f"data: {json.dumps({'error': str(e)})}\n\n"

# Health check endpoint
@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "service": "multimodal-llm-platform-simple"}

# Test Ollama endpoint
@app.get("/test-ollama")
async def test_ollama():
    """Test connection to local Ollama"""
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.get("http://localhost:11434/api/tags")
            if response.status_code == 200:
                models = response.json()
                return {
                    "status": "success",
                    "ollama_connected": True,
                    "available_models": models.get("models", [])
                }
            else:
                return {
                    "status": "error", 
                    "ollama_connected": False,
                    "error": f"HTTP {response.status_code}"
                }
    except Exception as e:
        return {
            "status": "error",
            "ollama_connected": False, 
            "error": str(e)
        }

# Chat with local Ollama
@app.post("/api/v1/chat/completions")
@performance_monitor("chat_completion")
@error_handler("autopicker.api.chat")
async def chat_completion(
    request: ChatCompletionRequest,
    user=Depends(get_optional_user)
):
    """Chat completion using local Ollama with security validation and logging"""
    start_time = time.time()
    try:
        # Sanitize input messages
        sanitized_messages = []
        for msg in request.messages:
            sanitized_content = security_manager.sanitize_input(msg.content, max_length=50000)
            sanitized_messages.append({"role": msg.role, "content": sanitized_content})
        
        # Convert to Ollama format
        payload = {
            "model": "llama3.2:1b",  # Use the actual model name
            "messages": sanitized_messages,
            "stream": request.stream
        }
        
        # Return streaming response if requested
        if request.stream:
            logger.info(f"Streaming chat completion request: {selected_model}")
            messages = [{"role": msg.role, "content": msg.content} for msg in request.messages]
            return StreamingResponse(
                stream_enhanced_response(selected_model, messages, temperature=request.temperature, max_tokens=request.max_tokens),
                media_type="text/plain",
                headers={
                    "Cache-Control": "no-cache",
                    "Connection": "keep-alive",
                    "Content-Type": "text/plain; charset=utf-8"
                }
            )
        
        logger.info(f"Sending request using enhanced router: {selected_model}")
        
        # Use enhanced router for API call
        messages = [{"role": msg.role, "content": msg.content} for msg in request.messages]
        openai_response = await enhanced_router.make_api_call(
            selected_model, 
            messages, 
            stream=False,
            temperature=request.temperature,
            max_tokens=request.max_tokens
        )
            
        logger.info(f"Successfully processed request with enhanced router")
        return openai_response
            
    except httpx.RequestError as e:
        logger.error(f"Request error: {e}")
        raise HTTPException(
            status_code=503,
            detail="Ollama is not available. Please ensure it's running."
        )
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Multimodal chat endpoint
@app.post("/api/v1/chat/multimodal")
async def multimodal_chat(request: MultimodalRequest):
    """Chat completion with file support and intelligent token management"""
    try:
        # Build the context with file contents
        context_messages = []
        
        # Collect file information and content for analysis
        file_info = []
        file_contents = []
        
        # Add file contents to context if provided
        if request.file_ids:
            for file_id in request.file_ids:
                # Find the file
                file_found = False
                for file_path in UPLOAD_DIR.iterdir():
                    if file_path.stem == file_id:
                        file_content = get_file_content(file_path)
                        file_type = file_content.get("type", "unknown")
                        
                        # Collect file info for routing
                        file_info.append({
                            "type": file_type,
                            "size": file_path.stat().st_size,
                            "name": file_path.name
                        })
                        
                        # Collect file content for token analysis
                        if file_type in ["pdf", "docx", "text", "excel"]:
                            file_contents.append({
                                "filename": file_path.name,
                                "type": file_type,
                                "content": file_content.get('content', 'No content available')
                            })
                        elif file_type == "image":
                            file_contents.append({
                                "filename": file_path.name,
                                "type": file_type,
                                "content": f"Image file: {file_content.get('description', 'Image file')}"
                            })
                        else:
                            file_contents.append({
                                "filename": file_path.name,
                                "type": file_type,
                                "content": f"File type {file_type} uploaded but not fully processed"
                            })
                        
                        file_found = True
                        break
                
                if not file_found:
                    file_contents.append({
                        "filename": f"file_id_{file_id}",
                        "type": "missing",
                        "content": f"File ID {file_id} not found"
                    })
        
        # Smart model selection (need this before token analysis)
        selected_model = enhanced_router.select_best_model(request, file_info)
        
        # Analyze token usage and determine if chunking is needed
        user_message_content = " ".join([msg.content for msg in request.messages])
        token_analysis = token_manager.analyze_content_for_chunking(
            file_contents=file_contents,
            model_id=selected_model,
            user_prompt=user_message_content
        )
        
        logger.info(f"Token analysis for {selected_model}: {token_analysis['total_estimated_tokens']} tokens, chunking_recommended: {token_analysis['chunking_recommended']}")
        
        # Handle chunking if needed
        if token_analysis['exceeds_limit'] or token_analysis['chunking_recommended']:
            logger.warning(f"Token limit approaching/exceeded for {selected_model}. Implementing intelligent summarization.")
            
            # Extract context keywords from user messages for better summarization
            context_keywords = []
            for msg in request.messages:
                # Simple keyword extraction from user query
                words = re.findall(r'\b[a-zA-Z]{4,}\b', msg.content.lower())
                context_keywords.extend([w for w in words if w not in ['this', 'that', 'with', 'from', 'they', 'were', 'been', 'have', 'will']])
            
            # Remove duplicates and limit
            context_keywords = list(set(context_keywords))[:10]
            
            # Use content summarization to fit within budget
            summaries = content_summarizer.batch_summarize_files(
                file_contents=file_contents,
                total_target_tokens=token_analysis['budget'].file_content,
                model_family=token_analysis['model_family'],
                context_keywords=context_keywords if context_keywords else None
            )
            
            file_context = "=== Attached Files (Intelligently Summarized) ===\n"
            total_compression = 0
            total_original = 0
            
            for i, summary in enumerate(summaries):
                filename = file_contents[i]['filename'] if i < len(file_contents) else 'unknown'
                file_context += f"\nFile: {filename}\n"
                file_context += f"Summary ({summary.strategy_used}, {summary.metadata.get('compression_percentage', 0)}% compressed):\n"
                file_context += f"{summary.summarized_content}\n"
                
                if summary.key_points:
                    file_context += f"Key Points: {'; '.join(summary.key_points[:3])}\n"
                
                file_context += "\n"
                total_compression += summary.summarized_tokens
                total_original += summary.original_tokens
            
            overall_compression = round((1 - total_compression / total_original) * 100, 1) if total_original > 0 else 0
            file_context += f"[System Note: Content summarized from {total_original} to {total_compression} tokens ({overall_compression}% compression) to fit context window. Key information preserved.]\n"
        
        else:
            # Normal processing - all content fits
            file_context = "=== Attached Files ===\n"
            for file_content in file_contents:
                file_context += f"\nFile: {file_content['filename']}\n"
                file_context += f"Type: {file_content['type']}\n"
                file_context += f"Content:\n{file_content['content']}\n\n"
        
        # Build context messages
        if file_contents:
            context_messages.append({
                "role": "system", 
                "content": f"You are analyzing the following files. Use this information to answer the user's questions:\n\n{file_context}"
            })
        
        # Add the user's messages
        for msg in request.messages:
            context_messages.append({"role": msg.role, "content": msg.content})
        
        # Convert to Ollama format
        payload = {
            "model": selected_model,
            "messages": context_messages,
            "stream": request.stream
        }
        
        # Return streaming response if requested
        if request.stream:
            logger.info(f"Streaming multimodal request with {len(request.file_ids)} files using model {selected_model}")
            return StreamingResponse(
                stream_enhanced_response(selected_model, context_messages),
                media_type="text/plain",
                headers={
                    "Cache-Control": "no-cache",
                    "Connection": "keep-alive",
                    "Content-Type": "text/plain; charset=utf-8"
                }
            )
        
        logger.info(f"Sending multimodal request using enhanced router with {len(request.file_ids)} files using model {selected_model}")
        
        # Use enhanced router for API call
        openai_response = await enhanced_router.make_api_call(
            selected_model, 
            context_messages, 
            stream=False
        )
        
        # Add files_processed count and token info to response
        openai_response["files_processed"] = len(request.file_ids) if request.file_ids else 0
        openai_response["token_usage"] = {
            "estimated_input_tokens": token_analysis['total_estimated_tokens'],
            "chunking_applied": token_analysis['chunking_recommended'] or token_analysis['exceeds_limit'],
            "model_context_window": token_analysis['budget'].max_context,
            "file_tokens": token_analysis['file_tokens'],
            "budget_breakdown": {
                "system_prompt": token_analysis['budget'].system_prompt,
                "user_prompt": token_analysis['user_prompt_tokens'],
                "file_content": token_analysis['file_tokens'],
                "response_buffer": token_analysis['budget'].response_buffer
            }
        }
        
        logger.info(f"Successfully processed multimodal request with {len(request.file_ids)} files, {token_analysis['total_estimated_tokens']} tokens")
        return openai_response
            
    except httpx.RequestError as e:
        logger.error(f"Request error: {e}")
        raise HTTPException(
            status_code=503,
            detail="Ollama is not available. Please ensure it's running."
        )
    except Exception as e:
        logger.error(f"Unexpected error in multimodal chat: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Audio transcription endpoint
@app.post("/api/v1/transcribe/{file_id}")
async def transcribe_audio(file_id: str):
    """Transcribe an uploaded audio file"""
    try:
        # Find the audio file
        audio_file_path = None
        for file_path in UPLOAD_DIR.iterdir():
            if file_path.stem == file_id:
                # Check if it's an audio file
                if file_path.suffix.lower() in ['.mp3', '.wav', '.m4a', '.ogg', '.flac']:
                    audio_file_path = file_path
                    break
                else:
                    raise HTTPException(status_code=400, detail=f"File {file_id} is not an audio file")
        
        if not audio_file_path:
            raise HTTPException(status_code=404, detail=f"Audio file {file_id} not found")
        
        # Process the audio file
        result = await process_audio(audio_file_path)
        
        if "error" in result:
            raise HTTPException(status_code=500, detail=result["error"])
        
        logger.info(f"Successfully transcribed audio file: {audio_file_path.name}")
        return {
            "file_id": file_id,
            "filename": audio_file_path.name,
            "transcription": result.get("transcription", ""),
            "language": result.get("language", "unknown"),
            "duration": result.get("duration", 0),
            "type": "audio"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Transcription error: {e}")
        raise HTTPException(status_code=500, detail=f"Transcription failed: {str(e)}")

# Enhanced multimodal chat that can handle audio transcription
@app.post("/api/v1/chat/multimodal-audio")
async def multimodal_audio_chat(request: MultimodalRequest):
    """Chat completion with file support including audio transcription"""
    try:
        # Build the context with file contents
        context_messages = []
        
        # Collect file information for smart routing
        file_info = []
        
        # Add file contents to context if provided
        if request.file_ids:
            file_context = "=== Attached Files ===\n"
            for file_id in request.file_ids:
                # Find the file
                file_found = False
                for file_path in UPLOAD_DIR.iterdir():
                    if file_path.stem == file_id:
                        file_type = file_path.suffix.lower()
                        
                        # Handle audio files with transcription
                        if file_type in ['.mp3', '.wav', '.m4a', '.ogg', '.flac']:
                            logger.info(f"Processing audio file for chat: {file_path.name}")
                            
                            # Collect file info for routing
                            file_info.append({
                                "type": "audio",
                                "size": file_path.stat().st_size,
                                "name": file_path.name
                            })
                            
                            audio_result = await process_audio(file_path)
                            
                            if "error" in audio_result:
                                file_context += f"\nFile: {file_path.name}\nType: Audio\nNote: {audio_result.get('fallback', 'Audio processing failed')}\n\n"
                            else:
                                transcription = audio_result.get("transcription", "")
                                language = audio_result.get("language", "unknown")
                                file_context += f"\nFile: {file_path.name}\nType: Audio Transcription\nLanguage: {language}\nContent:\n{transcription}\n\n"
                        else:
                            # Collect file info for routing
                            file_content = get_file_content(file_path)
                            file_info.append({
                                "type": file_content.get("type", "unknown"),
                                "size": file_path.stat().st_size,
                                "name": file_path.name
                            })
                            
                            # Handle other file types normally
                            file_type_name = file_content.get("type", "unknown")
                            
                            if file_type_name in ["pdf", "docx", "text", "excel"]:
                                file_context += f"\nFile: {file_path.name}\nType: {file_type_name}\nContent:\n{file_content.get('content', 'No content available')}\n\n"
                            elif file_type_name == "image":
                                file_context += f"\nFile: {file_path.name}\nType: Image\nDescription: {file_content.get('description', 'Image file')}\n\n"
                            else:
                                file_context += f"\nFile: {file_path.name}\nType: {file_type_name}\nNote: This file type is not fully processed but was uploaded successfully.\n\n"
                        
                        file_found = True
                        break
                
                if not file_found:
                    file_context += f"\nFile ID {file_id} not found.\n"
            
            # Add file context as a system message
            context_messages.append({
                "role": "system", 
                "content": f"You are analyzing the following files. Use this information to answer the user's questions:\n\n{file_context}"
            })
        
        # Add the user's messages
        for msg in request.messages:
            context_messages.append({"role": msg.role, "content": msg.content})
        
        # Smart model selection
        selected_model = enhanced_router.select_best_model(request, file_info)
        
        # Convert to Ollama format
        payload = {
            "model": selected_model,
            "messages": context_messages,
            "stream": request.stream
        }
        
        # Return streaming response if requested
        if request.stream:
            logger.info(f"Streaming multimodal-audio request with {len(request.file_ids)} files using model {selected_model}")
            return StreamingResponse(
                stream_enhanced_response(selected_model, context_messages),
                media_type="text/plain",
                headers={
                    "Cache-Control": "no-cache",
                    "Connection": "keep-alive",
                    "Content-Type": "text/plain; charset=utf-8"
                }
            )
        
        logger.info(f"Sending multimodal-audio request using enhanced router with {len(request.file_ids)} files using model {selected_model}")
        
        # Use enhanced router for API call
        openai_response = await enhanced_router.make_api_call(
            selected_model, 
            context_messages, 
            stream=False
        )
        
        # Add files_processed count to response
        openai_response["files_processed"] = len(request.file_ids) if request.file_ids else 0
        
        logger.info(f"Successfully processed multimodal-audio request with {len(request.file_ids)} files")
        return openai_response
            
    except httpx.RequestError as e:
        logger.error(f"Request error: {e}")
        raise HTTPException(
            status_code=503,
            detail="Ollama is not available. Please ensure it's running."
        )
    except Exception as e:
        logger.error(f"Unexpected error in multimodal-audio chat: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Smart routing analysis endpoint
@app.post("/api/v1/analyze-complexity")
async def analyze_complexity(request: MultimodalRequest):
    """Analyze request complexity and show model routing decision"""
    try:
        # Collect file information
        file_info = []
        
        if request.file_ids:
            for file_id in request.file_ids:
                for file_path in UPLOAD_DIR.iterdir():
                    if file_path.stem == file_id:
                        if file_path.suffix.lower() in ['.mp3', '.wav', '.m4a', '.ogg', '.flac']:
                            file_info.append({
                                "type": "audio",
                                "size": file_path.stat().st_size,
                                "name": file_path.name
                            })
                        else:
                            file_content = get_file_content(file_path)
                            file_info.append({
                                "type": file_content.get("type", "unknown"),
                                "size": file_path.stat().st_size,
                                "name": file_path.name
                            })
                        break
        
        # Calculate complexity
        complexity_score = model_selector.calculate_complexity_score(request, file_info)
        selected_model = enhanced_router.select_best_model(request, file_info)
        
        # Message analysis
        total_message_length = sum(len(msg.content) for msg in request.messages)
        
        return {
            "complexity_analysis": {
                "complexity_score": complexity_score,
                "selected_model": selected_model,
                "reasoning": {
                    "total_message_length": total_message_length,
                    "file_count": len(file_info),
                    "file_types": [f["type"] for f in file_info],
                    "total_file_size": sum(f["size"] for f in file_info),
                    "has_multimodal_content": len(file_info) > 0,
                    "complexity_factors": {
                        "message_complexity": "high" if total_message_length > 2000 else "medium" if total_message_length > 500 else "low",
                        "file_complexity": "high" if any(f["type"] in ["audio", "image"] for f in file_info) else "medium" if any(f["type"] in ["pdf", "excel"] for f in file_info) else "low" if file_info else "none",
                        "size_complexity": "high" if any(f["size"] > 1000000 for f in file_info) else "medium" if any(f["size"] > 100000 for f in file_info) else "low"
                    }
                }
            },
            "available_models": model_selector.models,
            "routing_logic": "Complexity-based model selection with file type and size analysis"
        }
        
    except Exception as e:
        logger.error(f"Complexity analysis error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Token analysis endpoint
@app.post("/api/v1/analyze-tokens")
async def analyze_tokens(request: MultimodalRequest):
    """Analyze token usage for a request without making API call"""
    try:
        # Collect file contents for analysis
        file_contents = []
        
        if request.file_ids:
            for file_id in request.file_ids:
                for file_path in UPLOAD_DIR.iterdir():
                    if file_path.stem == file_id:
                        file_content = get_file_content(file_path)
                        file_type = file_content.get("type", "unknown")
                        
                        if file_type in ["pdf", "docx", "text", "excel"]:
                            file_contents.append({
                                "filename": file_path.name,
                                "type": file_type,
                                "content": file_content.get('content', 'No content available')
                            })
                        elif file_type == "image":
                            file_contents.append({
                                "filename": file_path.name,
                                "type": file_type,
                                "content": f"Image file: {file_content.get('description', 'Image file')}"
                            })
                        else:
                            file_contents.append({
                                "filename": file_path.name,
                                "type": file_type,
                                "content": f"File type {file_type} uploaded but not fully processed"
                            })
                        break
        
        # Smart model selection for analysis
        file_info = [{"type": fc["type"], "name": fc["filename"]} for fc in file_contents]
        selected_model = enhanced_router.select_best_model(request, file_info)
        
        # Token analysis
        user_message_content = " ".join([msg.content for msg in request.messages])
        token_analysis = token_manager.analyze_content_for_chunking(
            file_contents=file_contents,
            model_id=selected_model,
            user_prompt=user_message_content
        )
        
        # Chunking analysis if needed
        chunking_info = None
        if token_analysis['chunking_recommended'] or token_analysis['exceeds_limit']:
            chunked_files = token_manager.chunk_files_for_model(
                file_contents=file_contents,
                model_id=selected_model,
                strategy=ChunkingStrategy.SEMANTIC
            )
            
            chunking_info = {
                "total_files": len(file_contents),
                "chunked_files": len(chunked_files),
                "chunk_details": []
            }
            
            for i, file_chunks in enumerate(chunked_files):
                if file_chunks:
                    chunking_info["chunk_details"].append({
                        "filename": file_chunks[0].source_file,
                        "total_chunks": file_chunks[0].total_chunks,
                        "chunks": [
                            {
                                "index": chunk.chunk_index,
                                "tokens": chunk.tokens,
                                "characters": chunk.metadata.get("character_count", 0),
                                "type": chunk.chunk_type
                            } for chunk in file_chunks
                        ]
                    })
        
        return {
            "model_selected": selected_model,
            "token_analysis": {
                "total_estimated_tokens": token_analysis['total_estimated_tokens'],
                "exceeds_limit": token_analysis['exceeds_limit'],
                "chunking_recommended": token_analysis['chunking_recommended'],
                "model_context_window": token_analysis['budget'].max_context,
                "usage_percentage": round((token_analysis['total_estimated_tokens'] / token_analysis['budget'].max_context) * 100, 2),
                "budget_breakdown": {
                    "system_prompt": token_analysis['budget'].system_prompt,
                    "user_prompt": token_analysis['user_prompt_tokens'],
                    "web_search": token_analysis['web_search_tokens'],
                    "file_content": token_analysis['file_tokens'],
                    "response_buffer": token_analysis['budget'].response_buffer,
                    "available_for_files": token_analysis['budget'].file_content
                },
                "file_analysis": token_analysis['file_analysis']
            },
            "chunking_analysis": chunking_info,
            "recommendations": {
                "use_chunking": token_analysis['chunking_recommended'],
                "recommended_strategy": "semantic" if token_analysis['chunking_recommended'] else None,
                "alternative_models": [
                    model_name for model_name, context in token_manager.model_contexts.items() 
                    if context > token_analysis['budget'].max_context
                ][:3],  # Top 3 larger context models
                "estimated_cost": token_manager.estimate_response_cost(
                    token_analysis['total_estimated_tokens'], 
                    selected_model
                )
            }
        }
        
    except Exception as e:
        logger.error(f"Token analysis error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# File upload
@app.post("/api/v1/upload", response_model=FileUploadResponse)
async def upload_file(
    file: UploadFile = File(...),
    user=Depends(get_optional_user)
):
    """Upload and save files with content processing and security validation"""
    try:
        # Security validation
        if not file.filename:
            raise HTTPException(status_code=400, detail="Filename is required")
        
        # Read file content first for validation
        content = await file.read()
        file_size = len(content)
        
        # Validate file security
        is_valid, validation_message = security_manager.validate_file_upload(
            file.filename, file.content_type or "application/octet-stream", file_size
        )
        
        if not is_valid:
            log_security_event("file_upload_rejected", {
                "filename": file.filename,
                "content_type": file.content_type,
                "size": file_size,
                "reason": validation_message,
                "user": user.get("username") if user else "anonymous"
            }, "WARNING")
            raise HTTPException(status_code=400, detail=validation_message)
        
        # Generate secure filename
        file_id = str(uuid.uuid4())
        secure_name = secure_filename(file.filename)
        file_extension = Path(secure_name).suffix if secure_name else ""
        stored_filename = f"{file_id}{file_extension}"
        file_path = UPLOAD_DIR / stored_filename
        
        # Save file
        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(content)
        
        # Process file content
        file_content = get_file_content(file_path)
        content_preview = None
        file_type = file_content.get("type", "unknown")
        
        # Create preview based on file type
        if file_type in ["pdf", "docx", "text"]:
            content_preview = file_content.get("content", "")[:200] + "..." if len(file_content.get("content", "")) > 200 else file_content.get("content", "")
        elif file_type == "image":
            content_preview = file_content.get("description", "Image file")
        elif file_type == "excel":
            content_preview = file_content.get("content", "")[:200] + "..." if len(file_content.get("content", "")) > 200 else file_content.get("content", "")
        
        # Log successful upload
        log_security_event("file_upload_success", {
            "filename": file.filename,
            "stored_as": stored_filename,
            "content_type": file.content_type,
            "size": file_size,
            "file_type": file_type,
            "user": user.get("username") if user else "anonymous"
        })
        
        logger.info(f"File uploaded and processed: {file.filename} -> {stored_filename} ({file_size} bytes, type: {file_type})")
        
        return FileUploadResponse(
            id=file_id,
            filename=stored_filename,
            original_filename=file.filename or "unknown",
            size=file_size,
            mime_type=file.content_type or "application/octet-stream",
            content_preview=content_preview,
            file_type=file_type
        )
        
    except Exception as e:
        logger.error(f"File upload error: {e}")
        raise HTTPException(status_code=500, detail=f"File upload failed: {str(e)}")

# List files
@app.get("/api/v1/files")
async def list_files():
    """List all uploaded files"""
    try:
        files = []
        for file_path in UPLOAD_DIR.iterdir():
            if file_path.is_file():
                stat = file_path.stat()
                files.append({
                    "filename": file_path.name,
                    "size": stat.st_size,
                    "created_at": stat.st_ctime,
                    "path": str(file_path)
                })
        
        return {"files": files, "count": len(files)}
        
    except Exception as e:
        logger.error(f"List files error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# Get available models
@app.get("/api/v1/models")
@performance_optimizer.cache_result(ttl=300)  # Cache for 5 minutes
@performance_optimizer.measure_performance("get_models")
async def get_models():
    """Get list of available models from enhanced router""" 
    try:
        # Get models from enhanced router
        models_list = enhanced_router.get_available_models()
        
        # Convert to OpenAI format
        openai_models = []
        for model in models_list:
            openai_models.append({
                "id": model["id"],
                "object": "model",
                "created": 0,
                "owned_by": f"{model['provider']}-autopicker",
                "description": model["description"],
                "context_length": model["context_length"],
                "capabilities": model["capabilities"],
                "cost_per_1k_tokens": model["cost_per_1k_tokens"],
                "available": model["available"]
            })
        
        return {
            "object": "list",
            "data": openai_models,
            "total_models": len(openai_models),
            "available_models": len([m for m in openai_models if m["available"]]),
            "providers": list(set(m["owned_by"] for m in openai_models))
        }
        
    except Exception as e:
        logger.error(f"Error fetching enhanced models: {e}")
        
        # Return minimal fallback models on error
        return {
            "object": "list",
            "data": [
                {
                    "id": "llama3.2-local",
                    "object": "model", 
                    "owned_by": "ollama-autopicker",
                    "description": "Local fallback model",
                    "available": True
                }
            ]
        }

# Performance endpoints
@app.get("/api/v1/performance/metrics")
async def get_performance_metrics():
    """Get comprehensive performance metrics"""
    try:
        return {
            "status": "healthy",
            "performance_metrics": performance_optimizer.get_performance_summary(),
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        log_error(e, {"endpoint": "performance_metrics"})
        raise HTTPException(status_code=500, detail="Failed to get performance metrics")

@app.post("/api/v1/performance/load-test")
async def run_load_test(
    endpoint: str = "/health",
    concurrent_users: int = 5,
    requests_per_user: int = 10
):
    """Run a load test on a specific endpoint"""
    try:
        if concurrent_users > 20 or requests_per_user > 50:
            raise HTTPException(status_code=400, detail="Load test parameters too high for safety")
        
        load_tester = LoadTester("http://localhost:8001")
        results = await load_tester.test_endpoint(
            endpoint=endpoint,
            concurrent_users=concurrent_users,
            requests_per_user=requests_per_user
        )
        
        return {
            "status": "completed",
            "load_test_results": results,
            "timestamp": datetime.now().isoformat()
        }
    except HTTPException:
        raise
    except Exception as e:
        log_error(e, {"endpoint": "load_test"})
        raise HTTPException(status_code=500, detail=f"Load test failed: {str(e)}")

@app.post("/api/v1/performance/comprehensive-test")
async def run_comprehensive_load_test():
    """Run comprehensive load tests on all key endpoints"""
    try:
        load_tester = LoadTester("http://localhost:8001")
        results = await load_tester.comprehensive_load_test()
        
        return {
            "status": "completed",
            "comprehensive_results": results,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        log_error(e, {"endpoint": "comprehensive_load_test"})
        raise HTTPException(status_code=500, detail=f"Comprehensive load test failed: {str(e)}")

# Root endpoint
@app.get("/")
async def root():
    """Root endpoint with API information"""
    return {
        "service": "Multimodal LLM Platform API - Simple",
        "version": "1.0.0",
        "description": "Simplified API for VPS deployment with Ollama integration",
        "endpoints": {
            "health": "/health",
            "test_ollama": "/test-ollama",
            "chat": "/api/v1/chat/completions",
            "multimodal_chat": "/api/v1/chat/multimodal",
            "multimodal_audio_chat": "/api/v1/chat/multimodal-audio",
            "transcribe": "/api/v1/transcribe/{file_id}",
            "analyze_complexity": "/api/v1/analyze-complexity",
            "analyze_tokens": "/api/v1/analyze-tokens",
            "upload": "/api/v1/upload",
            "files": "/api/v1/files",
            "models": "/api/v1/models",
            "monitoring": "/api/v1/monitoring/health",
            "logging": "/api/v1/logging/status",
            "performance_metrics": "/api/v1/performance/metrics",
            "load_test": "/api/v1/performance/load-test",
            "comprehensive_test": "/api/v1/performance/comprehensive-test"
        },
        "smart_routing": {
            "enabled": True,
            "default_model": "auto",
            "complexity_factors": ["message_length", "file_types", "file_sizes", "multimodal_content"],
            "routing_logic": "Automatic model selection based on request complexity"
        },
        "streaming": {
            "enabled": True,
            "description": "Real-time streaming responses supported",
            "parameter": "Set 'stream': true in request body"
        },
        "supported_file_types": {
            "documents": ["pdf", "docx", "txt", "md"],
            "spreadsheets": ["xlsx", "xls", "csv"],
            "images": ["jpg", "jpeg", "png", "gif", "bmp", "webp"],
            "audio": ["mp3", "wav", "m4a", "ogg", "flac"],
            "code": ["py", "js", "html", "css", "json"]
        },
        "docs": "/docs",
        "redoc": "/redoc"
    }

# Logging status endpoint
@app.get("/api/v1/logging/status")
async def get_logging_status_endpoint():
    """Get comprehensive logging system status"""
    try:
        status = get_logging_status()
        return {
            "status": "healthy",
            "logging_system": status,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        log_error(e, {"endpoint": "logging_status"})
        raise HTTPException(status_code=500, detail="Failed to get logging status")

# Usage statistics for billing (Stripe integration)
@app.get("/api/v1/billing/usage")
async def get_usage_statistics(user=Depends(get_optional_user)):
    """Get usage statistics for billing integration with Stripe"""
    try:
        usage_stats = enhanced_router.get_usage_stats()
        
        return {
            "status": "success",
            "usage_statistics": usage_stats,
            "billing_info": {
                "total_cost": usage_stats["total_cost"],
                "total_requests": usage_stats["total_requests"],
                "total_tokens": usage_stats["total_tokens"],
                "cost_breakdown_by_model": {
                    model_id: {
                        "requests": stats["total_requests"],
                        "tokens": stats["total_tokens"],
                        "cost": stats["total_cost"]
                    }
                    for model_id, stats in usage_stats["usage_by_model"].items()
                }
            },
            "pricing_tier": enhanced_router.default_pricing_tier,
            "enterprise_apis_enabled": enhanced_router.enable_enterprise_apis,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Usage statistics error: {e}")
        raise HTTPException(status_code=500, detail="Failed to get usage statistics")

# Pricing tiers endpoint
@app.get("/api/v1/pricing/tiers")
async def get_pricing_tiers():
    """Get available pricing tiers and model costs"""
    try:
        models = enhanced_router.get_available_models()
        
        standard_models = [m for m in models if m.get("pricing_tier", "standard") == "standard"]
        enterprise_models = [m for m in models if m.get("pricing_tier") == "enterprise"]
        
        return {
            "pricing_tiers": {
                "standard": {
                    "name": "Standard Tier",
                    "description": "OpenRouter-powered models for easy setup",
                    "features": [
                        "Access to 10+ AI models",
                        "Automatic model selection",
                        "Easy setup with single API key",
                        "Pay-per-use pricing"
                    ],
                    "models": standard_models,
                    "setup_required": "OpenRouter API key"
                },
                "enterprise": {
                    "name": "Enterprise Tier", 
                    "description": "Direct provider APIs for better pricing",
                    "features": [
                        "Better margins on model costs",
                        "Direct provider relationships",
                        "Enterprise negotiated rates",
                        "Advanced usage tracking",
                        "Priority support"
                    ],
                    "models": enterprise_models,
                    "setup_required": "Direct provider API keys",
                    "enabled": enhanced_router.enable_enterprise_apis
                }
            },
            "current_tier": enhanced_router.default_pricing_tier,
            "cost_comparison": {
                "gpt_4o_standard": next((m["cost_per_1k_tokens"] for m in standard_models if "gpt-4o" in m["id"]), None),
                "gpt_4o_enterprise": next((m["enterprise_cost"] for m in enterprise_models if "gpt-4o" in m["id"]), None),
                "claude_sonnet_standard": next((m["cost_per_1k_tokens"] for m in standard_models if "claude-3.5-sonnet" in m["id"]), None),
                "claude_sonnet_enterprise": next((m["enterprise_cost"] for m in enterprise_models if "claude-3.5-sonnet" in m["id"]), None)
            }
        }
    except Exception as e:
        logger.error(f"Pricing tiers error: {e}")
        raise HTTPException(status_code=500, detail="Failed to get pricing information")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001, reload=True)